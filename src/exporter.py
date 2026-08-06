"""
exporter.py: Export clean_news records to CSV, JSONL, or Excel files.

Entry point for 'python main.py export'.

  - Supported formats : CSV, JSONL, Excel (xlsx) — all 3 implemented.
  - Filtering option  : --status all | summarized | unsummarized
  - Output directory  : output/exports/

exporter.py
    ├── export_to_csv(records, out_path)
    ├── export_to_jsonl(records, out_path)
    ├── export_to_excel(records, out_path)
    └── export_news(args)              ← CLI entry point
"""

import csv
import json
import logging
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .database import get_clean_news

logger = logging.getLogger(__name__)

EXPORTS_DIR = Path("output") / "exports"

# Columns to include in exports (order matters for CSV/Excel header row)
EXPORT_COLUMNS = [
    "id",
    "title",
    "url",
    "source",
    "published_at",
    "category",
    "status",
    "snippet",
    "content",
    "summary",
    "sentiment",
    "method",
    "query",
    "collected_at",
]


# ---------------------------------------------------------------------------
# Format exporters
# ---------------------------------------------------------------------------

def export_to_csv(records, out_path):
    """Write records to a UTF-8 CSV file.
    UTF-8 with BOM (utf-8-sig) so Excel opens w/o garbling non-ASCII characters.
    Args:
        records:  List of dicts from get_clean_news().
        out_path: pathlib.Path for the output file.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        # utf-8-sig adds BOM so Excel opens the file correctly
        writer = csv.DictWriter(f, fieldnames=EXPORT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(records)

    logger.info(f"exporter: CSV saved -> {out_path} ({len(records)} rows)")
    return out_path


def export_to_jsonl(records, out_path):
    """Write records to a JSONL file (one JSON object per line, all 13 columns included).
    Args:
        records:  List of dicts from get_clean_news().
        out_path: pathlib.Path for the output file.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with out_path.open("w", encoding="utf-8") as f:
        for record in records:
            row = {col: record.get(col) for col in EXPORT_COLUMNS}
            f.write(json.dumps(row, ensure_ascii=False) + "\n")

    logger.info(f"exporter: JSONL saved -> {out_path} ({len(records)} rows)")
    return out_path


def export_to_excel(records, out_path):
    """Write records to an Excel (.xlsx) file with a styled header row.
    (.xlsx via openpyxl - styled blue header row, auto-fit column widths (capped at 60))
    Args:
        records:  List of dicts from get_clean_news().
        out_path: pathlib.Path for the output file.
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "clean_news"

    # --- Header row styling ---
    header_font   = Font(bold=True, color="FFFFFF")
    header_fill   = PatternFill(fill_type="solid", fgColor="4C72B0")
    header_align  = Alignment(horizontal="center", vertical="center", wrap_text=False)

    ws.append(EXPORT_COLUMNS)
    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    # --- Data rows ---
    for record in records:
        ws.append([record.get(col) for col in EXPORT_COLUMNS])

    # --- Auto-fit column widths (heuristic) ---
    for col_cells in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col_cells
        )
        # Cap width to avoid extremely wide columns (e.g. content)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 60)

    wb.save(out_path)
    logger.info(f"exporter: Excel saved -> {out_path} ({len(records)} rows)")
    return out_path


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def export_news(args):
    """Entry point for 'python main.py export'.
    [resolves status filter] → [fetches records] → [dispatches to the right exporter] 
    → [saves to output/exports/]
    Args:
        args.format : 'csv' | 'jsonl' | 'xlsx'  (required)
        args.status : 'all' | 'summarized' | 'unsummarized'  (default='all')
    """
    fmt    = args.format
    status = args.status  # 'all' | 'summarized' | 'unsummarized'

    # --- Fetch records with optional status filter ---
    status_filter = None if status == "all" else status
    records = get_clean_news(status=status_filter)

    if not records:
        msg = (
            f"No clean articles found"
            + (f" with status='{status}'" if status != "all" else "")
            + ". Nothing to export."
        )
        print(f"[WARNING] {msg}")
        logger.warning(f"export: {msg}")
        return

    print(f"[INFO] Exporting {len(records)} article(s) "
          f"(status={status}) as {fmt.upper()}...")

    # --- Build output file path ---
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"news_{status}_{timestamp}.{fmt if fmt != 'xlsx' else 'xlsx'}"
    out_path  = EXPORTS_DIR / filename

    # --- Dispatch to correct exporter ---
    if fmt == "csv":
        export_to_csv(records, out_path)
    elif fmt == "jsonl":
        export_to_jsonl(records, out_path)
    elif fmt == "xlsx":
        export_to_excel(records, out_path)
    else:
        print(f"[ERROR] Unsupported format: {fmt}")
        logger.error(f"export: unsupported format '{fmt}'")
        return

    print(f"[INFO] Export complete: {out_path}")
